#################################Autor: Jorge Eliécer Loaiza Muñoz###################################################

#Este código registra en SOFIA plus.
#Los documentos deben estar en una hoja de cálculo tipo excel.

from funciones import extraer #importa la función extraer desde el archivo funciones.py
from operator import truediv
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl
from openpyxl.styles import PatternFill
from tkinter import *
from tkinter import ttk, Button
from tkinter import filedialog, messagebox
from multiprocessing import Process
from threading import Thread  # librería para ejecutar en paralelo
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.opera import OperaDriverManager
# from python_anticaptcha import ImageToTextTask
# import pytesseract
# from PIL import Image
import string
import random

#file_path = 'F:/SENA Comercio y Servicios 2022/cursos/curso7/docs/8.pdf' #ruta de la foto del documento de identificación
path_name = 'F:/PythonProjects/registrador/datos_registrar_SOFIA.xlsx' # libro con los datos a registrar
wb = openpyxl.load_workbook(path_name)# cargo el libro donde están los datos a registrar
sheet = wb.active# se ubica en la hoja activa del libro donde deben estar los datos a verificar
# captura el número de la fila que contiene el último dato, es decir, la cantidad de filas con datos
ultima_fila_con_datos=0
ultima_fila_con_datos = sheet.max_row
print("Cantidad de documentos =",ultima_fila_con_datos - 1)

# esta línea alista el color verde
fill_pattern_verde = PatternFill(patternType='solid', fgColor='0099CC00')
# esta línea alista el color rojo
fill_pattern_rojo = PatternFill(patternType='solid', fgColor='00FF0000')
#Esta línea alista el color amarillo
fill_pattern_amarillo = PatternFill(patternType='solid', fgColor='FFFF00')

driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())  # para google chrome
driver.maximize_window()  # maximiza la ventana del navegador

for i in range(1,ultima_fila_con_datos+1):
#driver = webdriver.Opera(executable_path=OperaDriverManager().install())#para Opera
    driver.get('http://oferta.senasofiaplus.edu.co/sofia-oferta/inicio-sofia-plus.html')  # abre la página de SOFIA
    time.sleep(0.5)#le coloco este retraso de un segundo porque en unas pruebas no estaba cerrando la ventana.Tal vez esto sucedía
              #por lo rápido que se ejecuta el código
    driver.find_element_by_xpath('//*[@id="area_trabajo_divNoticiaPrincipal"]/div/span[1]/img').click()  # cierra el aviso inicial
    driver.find_element_by_xpath('//*[@id="registro"]')  # hace click en el botón "Registrarse"
    driver.get('http://oferta.senasofiaplus.edu.co/sofia-oferta/registro.html')  # se ubica en la página del registro
# esta línea permite cambiar al frame donde está el formulario para verificar si se está registrado
    driver.switch_to.frame(driver.find_element_by_xpath('//*[@id="modal-content"]/iframe'))

# Esta línea ubica la lista desplegable de los tipo de documentos de identidad
    element = driver.find_element_by_xpath('//*[@id="s1"]/select')
    dropdown = Select(element)
    #dropdown.select_by_visible_text('Tarjeta de Identidad') # para tarjeta de identidad
    dropdown.select_by_visible_text('Cédula de Ciudadanía') # para cédula de ciudadanía
    #dropdown.select_by_visible_text('Cédula de Extranjeria') # para cédula de extranjería
    #dropdown.select_by_visible_text('PEP') # para PEP

# esta línea me ubica en el cuadro de texto donde se debe ingresar el documento y borra su contenido
    driver.find_element_by_xpath('//*[@id="validar"]/label[2]/div[2]/input').clear()
# esta línea me ubica en el cuadro de texto donde se debe ingresar el documento e ingresa un documento
    driver.find_element_by_xpath('//*[@id="validar"]/label[2]/div[2]/input').send_keys(sheet.cell(row=i + 1, column=1).value)#escribe el documento de identidad
# esta línea hace clic en el botón "validar"
    driver.find_element_by_xpath('//*[@id="validar"]/div[4]/button').click()
# Espera dos segundos antes de continuar con el código de abajo. Es necesario esperar porque el aviso emergente de
# ya registrado se puede tardar uno o dos segundos en aparecer.
    time.sleep(2)

# Xpath del aviso emergente que indica que el documento ya está resgistrado
    display = driver.find_element_by_xpath('//*[@id="msg13"]')
    if display.is_displayed()== True:  # si es igual a True es porque apareció el aviso emergente, lo que significa que el documento ya está registrado
        sheet.cell(i + 1, 1).fill = fill_pattern_rojo  # rellena de rojo la celda con el documento que ya está registrado
        wb.save(path_name)
        print("Documento ya registrado")
    else:
        time.sleep(0.5)
        driver.find_element_by_xpath('//*[@id="gwt-uid-70"]').click()#hace click en la casilla "Acepto los términos de uso y política de confidencialidad y entiendo las responsabilidades derivadas de estas. "
        driver.find_element_by_xpath('//*[@id="registro_paso_1"]/div[2]/div[2]/button[2]').click()#hace click en el botón "Continuar"
        driver.find_element_by_xpath('//*[@id="div_usuario_documento"]/label/div[2]/button').click()#hace click en el botón "Examinar"
        element = driver.find_element_by_xpath('//*[@id="selectDoc"]/select')#despliega la lista.
        dropdown = Select(element)
        dropdown.select_by_visible_text('DOCUMENTO IDENTIDAD - PDF - MAX: 2 MB')#coloca en la lista "DOCUMENTO IDENTIDAD - PDF - MAX: 2 MB"

    # esta línea carga el documento que está en formato pdf
        driver.find_element_by_xpath('//*[@id="contentPopupCargarDocumento"]/div[2]/form/input').send_keys(sheet.cell(row=i + 1, column=2).value)#carga el path del documento que está en pdf.
        #time.sleep(3)
        driver.find_element_by_xpath('//*[@id="contentPopupCargarDocumento"]/div[3]/div/button[2]').click()# hace click en el botón "Guardar"
        #time.sleep(60)#esta línea es para darle tiempo al mensaje emergente de que aparezca

###########################################################################################################################################
### En ocasiones el aviso de carga exitosa del documento se demora en aparecer entonces para evitar esperar tanto tiempo con un time.sleep       
### implemento este ciclo que espera a que el aviso aparezca###############################################################################
        h=0
        while h==0:
                try:
                        #display = driver.find_element_by_id("modalBox")#esta línea captura el mensaje emergente de carga del documento.
                        # if display.is_displayed() == True:  # si es igual a True es porque apareció el aviso emergente, lo que significa que el documento se cargó correctamente
                        time.sleep(2)
                        driver.find_element_by_xpath('//*[@id="modal-close"]').click()#cierra el aviso emergente anterior
                        print("Documento cargado correctamente")
                        h=1
                        time.sleep(2)
                        
                except:
                        h=0

#############################################################################################################################################

        element = driver.find_element_by_xpath('//*[@id="s2"]/select')#despliega la lista de los paises
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=3).value)#selecciona el país especificado.

        element = driver.find_element_by_xpath('//*[@id="s3"]/select')  # despliega la lista de los departamentos
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=4).value)  # selecciona el departamento especificado.

        time.sleep(0.5)
        element = driver.find_element_by_xpath('//*[@id="s4"]/select')  # despliega la lista de los municipios
        dropdown = Select(element)
        time.sleep(0.5)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=5).value) # selecciona el municipio especificado.


#########Bloque para la fecha de expedición del documento#############################
        element = driver.find_element_by_xpath('//*[@id="s1001"]/select[1]')#Despliega la lista de los días que son 31.
        dropdown = Select(element)
        dropdown.select_by_visible_text(str(sheet.cell(row=i + 1, column=6).value))  # selecciona el día que es un número entre 1 y 31
        #hay que convertir a tipo cadena con "str", de lo contrario arroja el error "TypeError: argument of type 'int' is not iterable"

        element = driver.find_element_by_xpath('//*[@id="s1001"]/select[2]')  # Despliega la lista de los meses que son 12.
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=7).value)  # selecciona el mes.

        element = driver.find_element_by_xpath('//*[@id="s1001"]/select[3]')  # Despliega la lista de los años.
        dropdown = Select(element)
        dropdown.select_by_visible_text(str(sheet.cell(row=i + 1, column=8).value))  # selecciona el año.
        # hay que convertir a tipo cadena con "str", de lo contrario arroja el error "TypeError: argument of type 'int' is not iterable"
        time.sleep(0.5)
##################Bloque para nombre y apellidos##########################################################
        driver.find_element_by_xpath('//*[@id="div_usuario_nombre"]/label/div[2]/input').clear()#se ubica en el cuadro de texto del nombre y lo limpia
        driver.find_element_by_xpath('//*[@id="div_usuario_nombre"]/label/div[2]/input').send_keys(sheet.cell(row=i + 1, column=9).value)#ESCRIBE EL NOMBRE
        driver.find_element_by_xpath('//*[@id="div_usuario_apellido1"]/label/div[2]/input').clear()#se ubica en el cuadro de texto del primer apellido y lo limpia
        driver.find_element_by_xpath('//*[@id="div_usuario_apellido1"]/label/div[2]/input').send_keys(sheet.cell(row=i + 1, column=10).value)  # ESCRIBE EL PRIMER APELLIDO
        driver.find_element_by_xpath('//*[@id="div_usuario_apellido2"]/label/div[2]/input').clear()#se ubica en el cuadro de texto del SEGUNDO apellido y lo limpia
        driver.find_element_by_xpath('//*[@id="div_usuario_apellido2"]/label/div[2]/input').send_keys(sheet.cell(row=i + 1, column=11).value)  # ESCRIBE EL SEGUNDO APELLIDO

###################Género##########################################################################################
        element = driver.find_element_by_xpath('//*[@id="s5"]/select')
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=12).value)  # selecciona el género.

###################Fecha de nacimiento##########################################################################################
        element = driver.find_element_by_xpath('//*[@id="registro_paso_2"]/div[2]/div[2]/span[1]/span/select[1]')
        dropdown = Select(element)
        dropdown.select_by_visible_text(str(sheet.cell(row=i + 1, column=13).value))#Día

        element = driver.find_element_by_xpath('//*[@id="registro_paso_2"]/div[2]/div[2]/span[1]/span/select[2]')
        dropdown = Select(element)
        dropdown.select_by_visible_text(str(sheet.cell(row=i + 1, column=14).value))  #Mes

        element = driver.find_element_by_xpath('//*[@id="registro_paso_2"]/div[2]/div[2]/span[1]/span/select[3]')
        dropdown = Select(element)
        dropdown.select_by_visible_text(str(sheet.cell(row=i + 1, column=15).value))  #Año

###################País de nacimiento##########################################################################################
        element = driver.find_element_by_xpath('//*[@id="registro_paso_2"]/div[2]/div[2]/span[2]/select')#ubica los paises
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=16).value)  # Ingresa el país de nacimiento

###########################Departamento de nacimiento################################################
        time.sleep(0.5)
        element = driver.find_element_by_xpath('//*[@id="div_usuario_nac_dept"]/label/span/select')  # ubica los departamentos
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=17).value)  # Ingresa el departamento

###########################Municipio de nacimiento################################################
        time.sleep(0.4)
        element = driver.find_element_by_xpath('//*[@id="div_usuario_nac_mun"]/label/span/select')  # ubica los municipios
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=18).value)  # Ingresa el municipio

###########################estrato socioeconómico################################################
        time.sleep(0.5)
        element = driver.find_element_by_xpath('//*[@id="div_usuario_estrato"]/label/span[2]/select')  # ubica los estratos
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=19).value)  # Ingresa el estrato

###########################Botón "Continuar"################################################
        driver.find_element_by_xpath('//*[@id="registro_paso_2"]/div[2]/div[4]/button[2]').click()

###########################Pais de residencia"################################################
        element = driver.find_element_by_xpath('//*[@id="div_usuario_res_pais"]/label/span/select')  # ubica los paises
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=20).value)  # Ingresa el país de residencia

###########################Departamento de residencia"################################################
        time.sleep(1)
        element = driver.find_element_by_xpath('//*[@id="div_usuario_res_dept"]/label/span/select')  # ubica los departamentos
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=21).value)  # Ingresa el departamento de residencia
        time.sleep(1)
###########################Municipio de residencia"################################################
        time.sleep(1)
        element = driver.find_element_by_xpath('//*[@id="div_usuario_res_mun"]/label/span/select')  # ubica los municipios
        dropdown = Select(element)
        dropdown.select_by_visible_text(sheet.cell(row=i + 1, column=22).value)  # Ingresa el municipio de residencia

###########################Contacto personal correo electrónico"################################################
        time.sleep(0.7)
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[6]/input').clear()#limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[6]/input').send_keys(sheet.cell(row=i + 1, column=23).value)#ingresa el correo
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[8]/input').clear()#limpia el segundo cuadro de texto del correo(confirmación)
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[8]/input').send_keys(sheet.cell(row=i + 1, column=23).value)#Ingresa de nuevo el mismo correo.

###########################Contacto personal teléfono fijo"################################################
        time.sleep(0.5)
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[10]/div[3]/input').clear()  # limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[10]/div[3]/input').send_keys(sheet.cell(row=i + 1, column=24).value)  # limpia el cuadro de texto
                                     
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[10]/div[3]/input').clear() #limpia el cuadro de texto                         
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/div[10]/div[3]/input').send_keys(sheet.cell(row=i + 1, column=25).value)
################################################# prefijo Teléfono móvil#########################################################

        time.sleep(0.5)
        element = driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/span/select')  # ubica los prefijos
        dropdown = Select(element)
        dropdown.select_by_visible_text(str(sheet.cell(row=i + 1, column=26).value))  # Ingresa el prefijo

##################################################Número teléfono móvil######################################################
        time.sleep(0.5)
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/span/input').clear()#limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[1]/span/input').send_keys(sheet.cell(row=i + 1, column=27).value)#ingresa el número del móvil

###############Datos de un familiar o acudiente mayor de edad (nombres y apellidos)################################################
        time.sleep(0.5)
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[2]/input').clear()  # limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[2]/input').send_keys(sheet.cell(row=i + 1, column=28).value)  #Ingresa el nombre del acudiente

###########################Teléfono fijo del acudiente (prefijo)"################################################
        time.sleep(0.5)
        # driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[5]/div[3]/input').clear()  # limpia el cuadro de texto
        # driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[5]/div[3]/input').send_keys(sheet.cell(row=i + 1, column=29).value)  # ingresa el teléfono fijo del acudiente
                                     
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[5]/div[2]/span/input').clear() # limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[5]/div[2]/span/input').send_keys(sheet.cell(row=i + 1, column=29).value)

###########################Teléfono fijo del acudiente"#######################################################################
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[5]/div[3]/input').clear() #limpia el contenido
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[2]/div[5]/div[3]/input').send_keys(sheet.cell(row=i + 1, column=30).value)

##########################################Click en botón "continuar"###################################################################
        time.sleep(0.5)
        driver.find_element_by_xpath('//*[@id="registro_paso_3"]/div[2]/div[4]/button[2]').click()#hace click en el botón "Continuar".

########################### Contraseña de SOFIA"################################################
        time.sleep(0.5)
        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[1]/div[2]/input').clear()  # limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[1]/div[2]/input').send_keys('Sena12345')  # INTRODUCE EL TEXTO QUE SERÁ LA CONTRASEÑA
        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[1]/div[4]/input').clear()  # limpia el cuadro de texto
        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[1]/div[4]/input').send_keys('Sena12345')  # INTRODUCE EL TEXTO QUE SERÁ LA CONTRASEÑA

        
        driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').clear()
        driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').send_keys("")
        time.sleep(12)# este retraso es para tener tiempo de ingresar el captcha
                                      
        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[5]/button[2]').click()#hace click en el botón "Terminar"
        time.sleep(5)

        mensaje = driver.find_element_by_xpath('//*[@id="modal-content"]')#captura el mensaje obtenido despupes de hacer click en "Terminar"
        print(mensaje.text)
        mensaje_corto = extraer(mensaje.text)
        print (mensaje_corto)
        time.sleep(5)
        
        #El código de seguridad digitado por usted no coincide con el de la imágen.
        if mensaje_corto == "El código":

                h=0
                while h==0:
                        driver.find_element_by_xpath('//*[@id="modal-close"]').click() #cierra el aviso de captcha incorrecto
                        driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').clear()
                        driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').send_keys("")
                        time.sleep(5)
                        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[5]/button[2]').click()#hace click en el botón "Terminar"
                        time.sleep(5)
                        mensaje = driver.find_element_by_xpath('//*[@id="modal-content"]')
                        mensaje_corto = extraer(mensaje.text)

                        if mensaje_corto == "El código":
                                h=0
                        else:
                                #mensaje = driver.find_element_by_xpath('//*[@id="modal-content"]/div/div/div')
                                h=1

                if mensaje_corto == "Lo sentim":
                        time.sleep(1)
                        sheet.cell(i + 1,1).fill = fill_pattern_amarillo  # rellena de amarillo la celda con el documento en cuestión.
                        wb.save(path_name)
                        print("----> Después de realizada la validación de los datos registrados se encuentra que existe un usuario con la misma información ",sheet.cell(row=i + 1, column=1).value)
                
                else:
                        sheet.cell(i + 1, 1).fill = fill_pattern_verde  # rellena de verde la celda con el documento que acaba de ser registrado
                        wb.save(path_name)
                        print("----> Registro exitoso")
                        time.sleep(1)    

        elif mensaje_corto == "Lo sentim" :
                time.sleep(1)
                sheet.cell(i + 1,1).fill = fill_pattern_amarillo  # rellena de amarillo la celda con el documento en cuestión.
                wb.save(path_name)
                print("----> Después de realizada la validación de los datos registrados se encuentra que existe un usuario con la misma información ",sheet.cell(row=i + 1, column=1).value)

        else:
                sheet.cell(i + 1, 1).fill = fill_pattern_verde  # rellena de verde la celda con el documento que acaba de ser registrado
                wb.save(path_name)
                print("----> Registro exitoso")
                time.sleep(1)         

                
        #if mensaje.text == " Después de realizada la validación de los datos registrados se encuentra que existe un usuario con la misma información, por favor ingrese utilizando sus credenciales de usuario. Recuerde que si no recuerda su contraseña también la puede restablecer."
        #if mensaje.text == "Perfecto!!,  Su usuario ha sido creado correctamente.  A su cuenta de correo hemos enviado un mensaje confirmando los datos de su registro.  Bienvenido(a) a Sofia Plus. "
			


        #driver.find_element_by_xpath('//*[@id="modal-content"]/div/div/div/strong') #mensaje que indica que el captcha es incorrecto capturado por xpath
        #display = driver.find_element_by_class_name('error') #mensaje que indica que el captcha es incorrecto capturado por class
        #time.sleep(1)
        #if display.is_displayed == True:

#         h=0
#         while h==0:  
#                 try:
#                 # //*[@id="modal-content"]/div/div
#                 # //*[@id="modalBox"]
#                 # //*[@id="modal-content"]/div/div/divdgn
#                 # //*[@id="modal-close"]
#                 #//*[@id="modal-content"]/div/div/div
# # //*[@id="modalBox"]
# # //*[@id="modal-content"]/div/div
# # //*[@id="modal-content"]/div/div
#                         display=driver.find_element_by_xpath('//*[@id="modal-content"]/div/div/div/strong')#mensaje que indica que el captcha es incorrecto capturado por xpath
#                         if display.is_displayed == True:
#                                 driver.find_element_by_id('modal-close').click()#cierra el aviso de captcha incorrecto
#                                 #driver.find_element_by_xpath('//*[@id="modal-close"]').click() #cierra el aviso de captcha incorrecto                      
#                                 #driver.find_element_by_class_name("close").click() #cierra el aviso de captcha incorrecto
#                                 driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').clear()
#                                 driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').send_keys("")
#                                 time.sleep(10)
#                                 driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[5]/button[2]').click()#hace click en el botón "Terminar"
#                                 time.sleep(2)
#                                 #//*[@id="modal-content"]/div/div/div/strong
                                                
#                 except:
#                         h=1
                                                      
                        
#         try:      
#                         driver.find_element_by_xpath('//*[@id="modal-content"]/div/div/div')  # esta línea captura el mensaje emergente que informa que hay otro usuario con la misma información
#                         time.sleep(1)
#                         #driver.find_element_by_xpath('//*[@id="modal-close"]').click()  # cierra el aviso emergente anterior
#                         sheet.cell(i + 1,1).fill = fill_pattern_amarillo  # rellena de amarillo la celda con el documento en cuestión.
#                         wb.save(path_name)
#                         print("----> Después de realizada la validación de los datos registrados se encuentra que existe un usuario con la misma información ",sheet.cell(row=i + 1, column=1).value)
                        
#         except:
#                         sheet.cell(i + 1, 1).fill = fill_pattern_verde  # rellena de verde la celda con el documento que acaba de ser registrado
#                         wb.save(path_name)
#                         print("----> Registro exitoso")
#                         time.sleep(1)     

                       
        
                      
        
        

                # display = driver.find_element_by_xpath('//*[@id="modal-content"]/div/div/div')  # esta línea captura el mensaje emergente que informa que hay otro usuario con la misma información
                # time.sleep(1)
        
                # if display.is_displayed() == True:  # si es igual a True es porque apareció el aviso emergente, lo que significa que hay otro usuario con esa información
                #         #driver.find_element_by_xpath('//*[@id="modal-close"]').click()  # cierra el aviso emergente anterior
                #         sheet.cell(i + 1,1).fill = fill_pattern_amarillo  # rellena de amarillo la celda con el documento en cuestión.
                #         wb.save(path_name)
                #         print("Después de realizada la validación de los datos registrados se encuentra que existe un usuario con la misma información ",sheet.cell(row=i + 1, column=1).value)

        # cierra el aviso inicial
        #Cuando se da clic en el botón "Terminar" y el registro es satisfactorio, aparece la misma ventana o aviso emergente
        #que aparece al principio del proceso de registro. Una vez cerrada dicha ventana, aparece el mensaje de "Su usuario ha sido creado
        #satisfactoriamente"
                # else:
                #         time.sleep(1)
                #         #driver.find_element_by_xpath('//*[@id="area_trabajo_divNoticiaPrincipal"]/div/span[1]/img').click() # cierra el aviso inicial
                #         #time.sleep(1)
                #         #display = driver.find_element_by_xpath('//*[@id="modal-content"]/div/div')  # esta línea captura el mensaje emergente de  "Su usuario ha sido creado satisfactoriamente"
                #         #if display.is_displayed() == True:  # si es igual a True es porque apareció el aviso emergente, lo que significa que el registro fue satisfactorio.
                #         sheet.cell(i + 1, 1).fill = fill_pattern_verde  # rellena de verde la celda con el documento que acaba de ser registrado
                #         wb.save(path_name)
                #         print("Registro exitoso")
                #         time.sleep(1)



    #captcha = driver.find_element_by_xpath('//*[@id="div_captcha_code"]/img')#ubica el captcha
    #captcha_image = captcha.screenshot_as_png #toma una captura del captcha

##############################....................Captcha..................#############################################
##############...........Este bloque de código trata de resolver el captcha por fuerza bruta.............###############
'''
    h=0
    lista=list(string.ascii_lowercase)#esta línea genera las letras del alfabeto
    for i in range(10):#este ciclo agrega los números del 0 al 9 a la lista que contiene el alfabeto
        lista.append(i)

    while h==0:
        captcha=[]#lista donde se almacenará las letras que se pasarán como captcha
        for i in range(5):#este ciclo toma cinco caracteres de la lista y los guarda en la variable "captcha"
            j = random.randint(0, len(lista) - 1)
            captcha.append(lista[j])
        cadena="".join(map(str,captcha))#esta línea convierte a string cada caracter de la lista captcha y los concatena en una sola cadena de texto
        driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').clear()
        driver.find_element_by_xpath('//*[@id="div_captcha_code"]/input').send_keys(cadena)
        driver.find_element_by_xpath('//*[@id="registro_paso_4"]/div[2]/div[5]/button[2]').click()#hace click en el botón "Terminar"

        time.sleep(0.2)#le pongo este time.sleep para que le de diempo al mensaje de aparecer.
        display = driver.find_element_by_id("modalBox")# esta línea captura el mensaje emergente de cptcha erróneo.
        print("Encontró el mensaje emergente")
        if display.is_displayed() == True:  # si es igual a True es porque apareció el aviso emergente, lo que significa que el documento se cargó correctamente
            driver.find_element_by_xpath('//*[@id="modal-close"]').click()  # cierra el aviso emergente anterior
            print("Captcha incorrecto")
        else:
            print("Captcha correcto")
            h=1
'''
##################################################################################################################################################





print("vamos bien")



